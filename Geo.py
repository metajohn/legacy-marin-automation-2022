"""
Geo

Inputs the Geo information to a new bike

2023 TODO
#this function may require manual editing as it does not account for Frame Kits, go in and delete them manually
#Make sure formatting, year, and paths are correct

2022- TODO
check formatting on entry
"""

import openpyxl, os, time
from selenium import webdriver
from x_functions import x_click, x_delay_click, x_drop, x_login
from mFunctions import *

#
#
#
SaveState = define_save_state()
#
#
#


def geo(bike, yStart, yEnd):

    # waitForTasks()
    search(modelYear, bike)
    #Navigate to Geo
    x_delay_click(bikePD['Geo'])
    print('     Begin geo: ' + bike)

    """Geo Entry"""

    #TODO adjust table xpaths to account for new html
    #Add row on sitegraph
    for i in enumerate(orderList):
        x_click(bikePD['New_Geo'])
                
    #Add letters
    for i in enumerate(letterList):
        xKeys('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(i + 2) + ']/td[1]/textarea', letterList[i])
    #Add Categories
    for i in enumerate(orderList):
        xKeys('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(i + 2) + ']/td[2]/textarea', (geoSheet.cell(row=2, column=(orderList[i])).value))
    #Add Sizes
        #graphStart and sizecell are different only for readability, size indicators begin at column 3
    graphStart = 3
    sizeCell = graphStart
    for i in range(yStart, yEnd + 1):
        if((geoSheet.cell(row=i, column=1).value) is not None):
            xKeys('//*[@id="fields-bikeGeometryTable"]/tbody/tr[1]/td[' + str(sizeCell) + ']/textarea', geoSheet.cell(row=i, column=1).value)
            sizeCell += 1
    #Add Geo Values
    
    #accounting for single row geometries
    if ((yEnd - yStart) == 0):
        for o in enumerate(orderList):
            if (o == 2) or (o == 4):
                #Adds angle sign to correct rows
                xKeys('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[3]/textarea', (str((geoSheet.cell(row=yStart, column=(orderList[o])).value)))+ '°')
            else:
                xKeys('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[3]/textarea', str((geoSheet.cell(row=yStart, column=(orderList[o])).value)))
    #accounting for multiline geometries
    else:
        yTotal = yEnd - yStart
        yMod2 = yStart
        for o in enumerate(orderList):
            yMod = yStart
            while (yMod  < yMod2 + yTotal):
                for n in range(yTotal + 1):
                    #Adds angle sign to correct rows
                    if (o == 2) or (o == 4):
                        xKeys('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[' + str(n + 3) + ']/textarea', (str((geoSheet.cell(row=yMod, column=(orderList[o])).value)))+ '°')
                        yMod += 1
                    else:
                        xKeys('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[' + str(n + 3) + ']/textarea', str((geoSheet.cell(row=yMod, column=(orderList[o])).value)))
                        yMod += 1



    SaveBike(bike, SaveState, False)



runtime()

"""
SpreadSheet Mining
"""


#ANNUAL update bike spreadsheet PATH
wb = openpyxl.load_workbook('Marin Geos 2023 RUNTIME_09_29_2021.xlsx')
print(wb, 'open')


geoSheet = wb['Geometry ' + modelYear]
print(geoSheet, 'acquired')


#orderlist takes the order of the Geo Sheet and applies it to the websites ordering
orderList =  (14,  15,   2,   3,   4,   5,   6,   7,   8,   12,  11,  10,  9, 13, 16, 17, 18 )
letterList = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M')


"""
Site Navigation
"""

backstage_login()

NavBikes()

"""
Function Calls
"""

#All Bike spreadsheet cell values entered below
"""
COMPLETED

geo('Alpine Trail E',65,68)

geo('Wildcat Trail 2',94,97)
"""
geo('San Quentin Frame Kit',124,127)