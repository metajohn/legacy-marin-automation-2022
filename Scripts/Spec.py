"""
Spec

Adds Spec to new bike pages

2023 - the spreadsheets are now organized with bikes on different pages 
instead of all in one page

#2023 - preferred TODOS
#TODO add pagename to the geo function and manually enter the page name for each function call
#in addition to the yStart and yEned

#2023 - Ideal TODOS
#TODO create a function that finds the yStart and yEnd programatically
#TODO find the page of the bike programatically before geo begins
"""

#TODO add pagename to the geo function and manually enter the page for each function call
#in addition to the yStart and yEned

import openpyxl, os, time
from selenium import webdriver
from x_functions import x_click, x_delay_click, x_drop,x_login
from mFunctions import *

#
#
#
SaveState = define_save_state()
#
#
#

#TODO add pagename to the geo function and manually enter the page name for each function call
#in addition to the yStart and yEned
def spec(bike, cellStart, cellEnd, pageName):

    specSheet = wb[pageName]
    print(specSheet, 'acquired')

    NavBikes()
#    waitForTasks()

    #creates lists for title and body info of bike spec from specSheet

    #list int necessary for iterating
    listInt = 0
    titleList = list()
    bodyList = list()

    for x in range(cellStart, (cellEnd + 1)):
        titleList.insert(listInt, (specSheet.cell(row=x, column=3).value))
        bodyList.insert(listInt, (specSheet.cell(row=x, column=4).value))
        listInt += 1

    """
    Begin Site Interface
    """


    search(modelYear, bike)

    #switch to spec tab
    x_delay_click(bikePD['Spec'])

    #click new for number of cells
    totalCell = cellEnd - cellStart
    for i in range(1, (totalCell + 2)):
        x_click(bikePD['New_Spec'])
        time.sleep(1)
        #title
        xKeys('//*[@id="fields-bikeSpecifications-blocks-new' + str(i) + '-fields-specTitle"]', titleList[i-1])
        #body
        xKeys('//*[@id="fields-bikeSpecifications-blocks-new' + str(i) + '-fields-specContent-field"]/div[2]/div/div/div[2]', bodyList[i-1])
        
    time.sleep(1)

    """
    Save Entry
    """
    SaveBike(bike, SaveState, False)


"""
Begin
"""

runtime()

"""
                SpreadSheet Mining
"""

#Load spreadsheet

#TODO locate spreadsheet correctly
wb = openpyxl.load_workbook('Marin 2023 Catalog Spec RUNTIME 03_17_2022.xlsx')


"""
                Site Navigation
"""


backstage_login()

NavBikes()


"""
            Call Functions
"""
#TODO add pagename to the geo function and manually enter the page name for each function call
#in addition to the yStart and yEnd
        

"""
COMPLETED

spec('Wildcat Trail 2',26,49,'Wildcat Trail')

spec('Alpine Trail E',62,91,'Alpine Trail E')
"""
spec('San Quentin Frame Kit',77,81,'San Quentin')

