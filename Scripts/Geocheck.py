"""
GeoCheck
Compares site - spreadsheet values for geometry

DOES NOT ENTER NEW VALUES OR SAVE, FOR THIS USE GeoUpdate.py
"""

"""
TODO

delete any form of saving, 
when geodifferent == True
    add bike to list

AT THE END
    print list

this should be done so that the process can be repeated with only the bikes that have differences to confirm and then run geoupdate
"""

import openpyxl, os, time
from selenium import webdriver
from bikeList_2023 import bikeList
from x_functions import x_click, x_delay_click, x_drop,x_login
from mFunctions import *

# Adds new Geo to bike 2022
    
def geo(bike, yStart, yEnd, isFrameKit):
    

    SaveState = define_save_state()
    differentGeo = False

    NavBikes()
    #waitForTasks()
    search(modelYear, bike)
    
    #Navigate to Geo
    x_delay_click(bikePD['Geo'])
    print(f'     Begin geo CHECK: {bike}\n\n')


    """
    Geo Entry
    """
    if (isFrameKit == True):
        orderList = (14,  15,   2,   3,   4,   5,   6,   7,   8,   12,  11,  10,  9)
    else: 
        orderList = (14,  15,   2,   3,   4,   5,   6,   7,   8,   12,  11,  10,  9, 13, 16, 17, 18 )


    #accounting for single row geometries
    if ((yEnd - yStart) == 0):
        for o in range(len(orderList)):
            if (o == 2) or (o == 4):
                #Adds angle sign to correct rows
                sheetValue = (str((geoSheet.cell(row=yStart, column=(orderList[o])).value))+ '°')
                siteValue = (str(x_get_text('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[3]/textarea')))
                if (siteValue != sheetValue):
                    differentGeo = True
                    print(f'Difference --  sheetValue: {sheetValue}, siteValue: {siteValue} -- siteGraphLocation: x: 3 y: {o + 2}')
                    """
                    Disabled for testing
                    xKeys('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[3]/textarea', (str((geoSheet.cell(row=yStart, column=(orderList[o])).value)))+ '°')
                    """
            else:
                sheetValue = (str((geoSheet.cell(row=yStart, column=(orderList[o])).value)))
                siteValue = str(x_get_text('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[3]/textarea'))
                if (siteValue != sheetValue):
                    differentGeo = True
                    print(f'Difference --  sheetValue: {sheetValue}, siteValue: {siteValue} -- siteGraphLocation: x: 3 y: {o + 2}')
                    """
                    Disabled for testing
                    xKeys('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[3]/textarea', str((geoSheet.cell(row=yStart, column=(orderList[o])).value)))
                    """
        if differentGeo == False:
            print('No differences detected in Geo')
    #accounting for multi row geometries
    else:
        yTotal = yEnd - yStart
        yMod2 = yStart
        for o in range(len(orderList)):
            yMod = yStart
            while (yMod  < yMod2 + yTotal):
                for n in range(yTotal + 1):
                    #Adds angle sign to correct rows
                    if (o == 2) or (o == 4):
                        sheetValue = (str((geoSheet.cell(row=yMod, column=(orderList[o])).value))+ '°')
                        siteValue = x_get_text('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[' + str(n + 3) + ']/textarea')
                        if (siteValue != sheetValue):
                            differentGeo = True
                            print(f'Difference --  sheetValue: {sheetValue}, siteValue: {siteValue} -- siteGraphLocation: x: {n + 3} y: {o + 2} -- sheetGraphLocation x: {o} y: {yMod}')
                            """
                            Disabled for testing
                            xKeys('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[' + str(n + 3) + ']/textarea', (str((geoSheet.cell(row=yMod, column=(orderList[o])).value)))+ '°')
                            """
                                                    
                        yMod += 1
                    else:
                        sheetValue = str((geoSheet.cell(row=yMod, column=(orderList[o])).value))
                        siteValue = x_get_text('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[' + str(n + 3) + ']/textarea')
                        if (siteValue != sheetValue):
                            differentGeo = True
                            print(f'Difference --  sheetValue: {sheetValue}, siteValue: {siteValue} -- siteGraphLocation: x: {n + 3} y: {o + 2} -- sheetGraphLocation x: {o} y: {yMod}')
                            """
                            Disabled for testing
                            xKeys('//*[@id="fields-bikeGeometryTable"]/tbody/tr[' + str(o + 2) + ']/td[' + str(n + 3) + ']/textarea', str((geoSheet.cell(row=yMod, column=(orderList[o])).value)))
                            """
                        yMod += 1
        if differentGeo == False:
            print('No differences detected in Geo')
        else:
            print(f'discrepancy detected, ADDED {bike} to discrepancy list')
            differentGeoList.append(f'geo(\'{bike}\',{yStart},{yEnd},{isFrameKit})')
            SaveState = define_save_state()



    print(f'\n\n     End geo CHECK: {bike}\n\n')
    SaveBike(bike, SaveState, False)
    



runtime()

differentGeoList = []

"""
SpreadSheet Mining
"""


#ANNUAL update bike spreadsheet PATH
wb = openpyxl.load_workbook('Marin Geos 2023 RUNTIME_09_29_2021.xlsx')
print(wb, 'open')
geoSheet = wb['Geometry ' + modelYear]
print(geoSheet, 'acquired')

#orderlist takes the order of the Geo Sheet and applies it to the websites ordering
orderList =  (14,  15,   2,   3,   4,   5,   6,   7,   8,   12,  11,  10,  9, 13, 16, 17, 18 )
letterList = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M')

"""
Site Navigation
"""

backstage_login()

NavBikes()

"""
Function Calls
"""

#All Bike spreadsheet cell values entered below
#TODO reverse these values????????????
"""
COMPLETED
"""
geo('Alpine Trail Carbon 2',41,44,False)
geo('Alpine Trail Carbon 1',47,50,False)
geo('Alpine Trail XR',53,56,False)
geo('Alpine Trail 7',59,62,False)
geo('Alpine Trail E2',65,68,False)
geo('Alpine Trail E1',65,68,False)
geo('Alpine Trail E',65,68,False)
geo('Alpine Trail Carbon 2 Frame Kit',41,44,True)
geo('Rift Zone 29" Carbon XR',10,13,False)
geo('Rift Zone 29" Carbon 2',4,7,False)
geo('Rift Zone 29" Carbon 1',4,7,False)
geo('Rift Zone 29" XR',16,19,False)
geo('Rift Zone 29" 2',16,19,False)
geo('Rift Zone 29" 1',16,19,False)
geo('Rift Zone 29" Carbon XR Frame Kit',10,13,True)
geo('Rift Zone 27.5" XR',35,38,False)
geo('Rift Zone 27.5" 2',29,32,False)
geo('Rift Zone 27.5" 1',22,26,False)
geo('San Quentin 3',124,127,False)
geo('San Quentin 2',124,127,False)
geo('San Quentin 1',118,121,False)
geo('San Quentin Frame Kit',124,127,True)
geo('El Roy',130,131,False)
geo('El Roy Frame Kit',130,131,True)
geo('Bobcat Trail 5',86,91,False)
geo('Bobcat Trail 4',86,91,False)
geo('Bobcat Trail 3',86,91,False)
geo('Bolinas Ridge 2',77,83,False)
geo('Bolinas Ridge 1',77,83,False)
geo('Wildcat Trail 3',100,103,False)
geo('Wildcat Trail 2',94,97,False)
geo('Wildcat Trail 1',94,97,False)
geo('Team Marin 2',106,109,False)
geo('Team Marin 1',106,109,False)
geo('Pine Mountain 2',112,115,False)
geo('Pine Mountain 1',112,115,False)
geo('Alcatraz',306,307,False)
geo('Alcatraz Frame Kit',306,307,True)
geo('Headlands 2',165,170,False)
geo('Headlands 1',165,170,False)
geo('Headlands Frame Kit',165,170,True)
geo('Gestalt XR',159,162,False)
geo('Gestalt X10',159,162,False)
geo('Gestalt 2',151,156,False)
geo('Gestalt 1',151,156,False)
geo('Gestalt',151,156,False)
geo('Nicasio 2',143,148,False)
geo('Nicasio+',134,140,False)
geo('Nicasio',134,140,False)
geo('Four Corners',181,185,False)
geo('Lombard 1',173,178,False)
geo('Fairfax 3',188,192,False)
geo('Fairfax 2',188,192,False)
geo('Fairfax ST 2',195,198,False)
geo('Fairfax 1',188,192,False)
geo('Fairfax ST 1',195,198,False)
geo('Presidio 3',201,205,False)
geo('Presidio 2',201,205,False)
geo('Presidio 1',201,205,False)
geo('Muirwoods',208,212,False)
geo('Larkspur 2',261,263,False)
geo('Larkspur 1',261,263,False)
geo('San Rafael DS2',215,219,False)
geo('San Rafael DS1',215,219,False)
geo('San Anselmo DS2',222,224,False)
geo('San Anselmo DS1',222,224,False)
geo('DSX FS',233,236,False)
geo('DSX 2',227,230,False)
geo('DSX 1',227,230,False)
geo('DSX',227,230,False)
geo('Kentfield 2',250,253,False)
geo('Kentfield ST 2',256,258,False)
geo('Kentfield 1',250,253,False)
geo('Kentfield ST 1',256,258,False)
geo('Stinson 2',266,269,False)
geo('Stinson ST 2',272,274,False)
geo('Stinson 1',266,269,False)
geo('Stinson ST 1',272,274,False)
geo('Sausalito E2',239,242,False)
geo('Sausalito ST E2',245,247,False)
geo('Sausalito E1',239,242,False)
geo('Sausalito ST E1',245,247,False)
geo('Rift Zone 26"',294,294,False)
geo('Rift Zone Jr 24\"',297,297,False)
geo('San Quentin 24\"',291,291,False)
geo('San Quentin 20\"',288,288,False)
geo('Bayview Trail 24\"',303,303,False)
geo('Hidden Canyon 20\"',300,300,False)

#print List of incorrect geo when complete with functions
print('/n Different Geos listed below')
for i in differentGeoList:
    print(i)