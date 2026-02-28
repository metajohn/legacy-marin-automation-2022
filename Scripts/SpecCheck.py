"""
SpecCheck
Compares site - spreadsheet values for spec

Unfinished

DOES NOT ENTER NEW VALUES OR SAVE, FOR THIS USE SpecUpdate.py
"""

"""
TODO
"""

import openpyxl, os, time
from selenium import webdriver
from x_functions import x_click, x_delay_click, x_drop,x_login, x_get_value, x_get_text
from mFunctions import *
    
def spec(bike, startCell, endCell, pageName):

    specSheet = wb[pageName]
    print(specSheet, 'acquired')

    SaveState = define_save_state()
    differentSpec = False

    NavBikes()
    #waitForTasks()
    search(modelYear, bike)
    
    #Navigate to Geo
    x_delay_click(bikePD['Spec'])

    print(f'     Begin spec CHECK: {bike}\n\n')
    #designed to record more interesting notes

    bikeHasProblems = False
    specRangeLength = endCell - startCell
    for i in range (1, specRangeLength+1):
        print('\n\n')

        #title
        siteTitle = x_get_value(f'/html/body/div[1]/div/div[3]/main/form/div[2]/div[1]/div/div[1]/div[3]/div/div[2]/div/div[1]/div[{i}]/div[4]/div/div[1]/div[2]/input')
        sheetTitle = str((specSheet.cell(row=(startCell + i - 1), column=3).value))
        print(siteTitle)
        print(sheetTitle)


        #value
        siteValue = x_get_text(f'/html/body/div[1]/div/div[3]/main/form/div[2]/div[1]/div/div[1]/div[3]/div/div[2]/div/div[1]/div[{i}]/div[4]/div/div[2]/div[2]/div/div/div[2]')
        sheetValue = str((specSheet.cell(row=(startCell + i - 1), column=4).value))
        print(siteValue)
        print(sheetValue)
        if(siteValue != sheetValue):
            differentSpec = True

            #differentiate between titles that are actually wrong and titles that are just differently entered
            siteTitle = siteTitle.lower()
            sheetTitle = sheetTitle.lower()
            if(siteTitle[0:3] != sheetTitle [0:3]):
                if (bikeHasProblems == False):
                    differentSpecNotes.append(f'\n++++++++++++++++++++++++++++\n{bike}\n\n')
                    bikeI = differentSpecNotes.index(f'\n++++++++++++++++++++++++++++\n{bike}\n\n')
                    bikeHasProblems = True
                differentSpecNotes[bikeI] = (differentSpecNotes[bikeI] + (f'\n SITE {siteTitle}\n  {siteValue} \n SHEET {sheetTitle}\n  {sheetValue} \n ------------------------'))

    """
    Spec Discrepancy
    """

    if (differentSpec == True):
        print('discrepancy detected, added to difference list')
        differentSpecList.append(f'spec(\'{bike}\',{startCell},{endCell},\'{pageName}\')')
    else:
        print('No differences detected in Spec')
        SaveState = define_save_state()



    print(f'\n\n     End spec CHECK: {bike}\n\n')
    SaveBike(bike, SaveState, False)
    



runtime()

"""
SpreadSheet Mining
"""


#TODO update bike spreadsheet PATH

wb = openpyxl.load_workbook('Marin 2023 Catalog Spec RUNTIME 03_17_2022.xlsx')
print(wb, 'open')

differentSpecList = []
differentSpecNotes = []

"""
Site Navigation
"""

backstage_login()

NavBikes()

"""
Function Calls
"""


"""
COMPLETED
"""

spec('Alpine Trail Carbon 2',2,28,'Alpine Trail')
spec('Alpine Trail Carbon 1',29,55,'Alpine Trail')
spec('Alpine Trail XR',56,82,'Alpine Trail')
spec('Alpine Trail 7',83,109,'Alpine Trail')
spec('Alpine Trail E2',2,31,'Alpine Trail E')
spec('Alpine Trail E1',32,61,'Alpine Trail E')
spec('Alpine Trail E',62,91,'Alpine Trail E')
spec('Alpine Trail Carbon 2 Frame Kit',110,115,'Alpine Trail')
spec('Rift Zone 29" Carbon XR',2,27,'Rift Zone (analog)')
spec('Rift Zone 29" Carbon 2',28,53,'Rift Zone (analog)')
spec('Rift Zone 29" Carbon 1',54,79,'Rift Zone (analog)')
spec('Rift Zone 29" XR',80,105,'Rift Zone (analog)')
spec('Rift Zone 29" 2',106,131,'Rift Zone (analog)')
spec('Rift Zone 29" 1',132,157,'Rift Zone (analog)')
spec('Rift Zone 29" Carbon XR Frame Kit',158,163,'Rift Zone (analog)')
spec('Rift Zone 27.5" XR',164,189,'Rift Zone (analog)')
spec('Rift Zone 27.5" 2',190,215,'Rift Zone (analog)')
spec('Rift Zone 27.5" 1',216,241,'Rift Zone (analog)')
spec('San Quentin 3',2,26,'San Quentin')
spec('San Quentin 2',27,51,'San Quentin')
spec('San Quentin 1',52,76,'San Quentin')
spec('San Quentin Frame Kit',77,81,'San Quentin')
spec('El Roy',2,26,'El Roy')
spec('El Roy Frame Kit',27,30,'El Roy')
spec('Bobcat Trail 5',2,25,'Bobcat Trail')
spec('Bobcat Trail 4',26,49,'Bobcat Trail')
spec('Bobcat Trail 3',50,73,'Bobcat Trail')
spec('Bolinas Ridge 2',2,25,'Bolinas Ridge')
spec('Bolinas Ridge 1',26,49,'Bolinas Ridge')
spec('Wildcat Trail 3',2,25,'Wildcat Trail')
spec('Wildcat Trail 2',26,49,'Wildcat Trail')
spec('Wildcat Trail 1',50,73,'Wildcat Trail')
spec('Team Marin 2',2,25,'Team Marin')
spec('Team Marin 1',26,49,'Team Marin')
spec('Pine Mountain 2',2,26,'Pine Mountain')
spec('Pine Mountain 1',27,50,'Pine Mountain')
spec('Alcatraz',2,27,'Alcatraz')
spec('Alcatraz Frame Kit',28,31,'Alcatraz')
spec('Headlands 2',2,26,'Headlands')
spec('Headlands 1',27,51,'Headlands')
spec('Headlands Frame Kit',2,5,'Headlands Frame Kit')
spec('Gestalt XR',2,26,'Gestalt')
spec('Gestalt X10',27,51,'Gestalt')
spec('Gestalt 2',52,76,'Gestalt')
spec('Gestalt 1',77,100,'Gestalt')
spec('Gestalt',101,124,'Gestalt')
spec('Nicasio 2',2,26,'Nicasio')
spec('Nicasio+',51,74,'Nicasio')
spec('Nicasio',27,50,'Nicasio')
spec('Four Corners',2,25,'Four Corners')
spec('Lombard 1',2,25,'Lombard')
spec('Fairfax 3',2,25,'Fairfax')
spec('Fairfax 2',26,49,'Fairfax')
spec('Fairfax ST 2',50,73,'Fairfax')
spec('Fairfax 1',74,97,'Fairfax')
spec('Fairfax ST 1',98,121,'Fairfax')
spec('Presidio 3',2,25,'Presidio')
spec('Presidio 2',26,49,'Presidio')
spec('Presidio 1',50,73,'Presidio')
spec('Muirwoods',2,25,'Muirwoods')
spec('Larkspur 2',2,25,'Larkspur')
spec('Larkspur 1',26,49,'Larkspur')
spec('San Rafael DS2',2,25,'San Rafael San Anselmo')
spec('San Rafael DS1',26,49,'San Rafael San Anselmo')
spec('San Anselmo DS2',50,73,'San Rafael San Anselmo')
spec('San Anselmo DS1',74,97,'San Rafael San Anselmo')
spec('DSX FS',77,101,'DSX')
spec('DSX 2',52,76,'DSX')
spec('DSX 1',27,51,'DSX')
spec('DSX',2,26,'DSX')
spec('Kentfield 2',2,25,'Kentfield')
spec('Kentfield ST 2',26,49,'Kentfield')
spec('Kentfield 1',50,73,'Kentfield')
spec('Kentfield ST 1',74,97,'Kentfield')
spec('Stinson 2',2,25,'Stinson')
spec('Stinson ST 2',26,49,'Stinson')
spec('Stinson 1',50,73,'Stinson')
spec('Stinson ST 1',74,97,'Stinson')
spec('Sausalito E2',2,28,'Sausalito')
spec('Sausalito ST E2',29,55,'Sausalito')
#spec('Sausalito E1',,,'Sausalito')
#spec('Sausalito ST E1',,,'Sausalito')
spec('Rift Zone 26"',267,292,'Rift Zone (analog)')
spec('Rift Zone Jr 24"',242,266,'Rift Zone (analog)')
spec('San Quentin 24"',2,25,'San Quentin Kids')
spec('San Quentin 20"',26,49,'San Quentin Kids')
spec('Bayview Trail 24"',2,25,'Kids')
spec('Hidden Canyon 20"',26,49,'Kids')


print('More interesting notes')

for i in differentSpecNotes:
    print(f'\n+++++++++++++++++++++++\n\n{i}')

print('FUNCTION BELOW FOR RUNNING specupdate.py\n\n')
#formatted discrepancy list for specUpdate
for i in differentSpecList:
    print(i)